'''
Business: Экспорт результатов опросов в PDF и Excel форматы
Args: event - dict с httpMethod, queryStringParameters (poll_id, format)
      context - object с attributes: request_id, function_name
Returns: HTTP response с файлом для скачивания
'''
import json
import os
import psycopg2
import io
import base64
from typing import Dict, Any
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference

def get_db_connection():
    dsn = os.environ.get('DATABASE_URL')
    return psycopg2.connect(dsn)

def generate_pdf(poll_data: Dict[str, Any], stats: list) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    story = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        textColor=colors.HexColor('#403E43')
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=12,
        textColor=colors.HexColor('#1EAEDB')
    )
    
    story.append(Paragraph('Результаты опроса', title_style))
    story.append(Spacer(1, 12))
    
    story.append(Paragraph(f'<b>Целевая аудитория:</b> {poll_data["target_audience"]}', styles['Normal']))
    story.append(Spacer(1, 12))
    
    story.append(Paragraph(f'<b>Вопрос:</b> {poll_data["question"]}', styles['Normal']))
    story.append(Spacer(1, 24))
    
    story.append(Paragraph('Статистика ответов', heading_style))
    story.append(Spacer(1, 12))
    
    total_votes = sum(stats)
    
    summary_data = [
        ['Показатель', 'Значение'],
        ['Всего голосов', str(total_votes)],
        ['Дата создания', poll_data['created_at'].split('T')[0] if poll_data.get('created_at') else 'N/A']
    ]
    
    summary_table = Table(summary_data, colWidths=[8*cm, 8*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1EAEDB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 24))
    
    story.append(Paragraph('Распределение голосов', heading_style))
    story.append(Spacer(1, 12))
    
    votes_data = [['Вариант ответа', 'Голосов', 'Процент']]
    
    for i, option in enumerate(poll_data['options']):
        count = stats[i]
        percentage = round((count / total_votes * 100) if total_votes > 0 else 0, 1)
        votes_data.append([option, str(count), f'{percentage}%'])
    
    votes_table = Table(votes_data, colWidths=[10*cm, 3*cm, 3*cm])
    votes_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#403E43')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(votes_table)
    story.append(Spacer(1, 24))
    
    story.append(Paragraph(f'Отчет сгенерирован: {datetime.now().strftime("%d.%m.%Y %H:%M")}', styles['Italic']))
    
    doc.build(story)
    
    pdf_bytes = buffer.getvalue()
    buffer.close()
    
    return pdf_bytes

def generate_excel(poll_data: Dict[str, Any], stats: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Результаты опроса'
    
    header_fill = PatternFill(start_color='1EAEDB', end_color='1EAEDB', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=14)
    
    ws['A1'] = 'РЕЗУЛЬТАТЫ ОПРОСА'
    ws['A1'].font = Font(bold=True, size=18, color='403E43')
    ws.merge_cells('A1:D1')
    
    ws['A3'] = 'Целевая аудитория:'
    ws['B3'] = poll_data['target_audience']
    ws['A3'].font = Font(bold=True)
    
    ws['A4'] = 'Вопрос:'
    ws['B4'] = poll_data['question']
    ws['A4'].font = Font(bold=True)
    ws.merge_cells('B4:D4')
    
    ws['A6'] = 'Общая статистика'
    ws['A6'].font = Font(bold=True, size=12)
    
    total_votes = sum(stats)
    
    ws['A7'] = 'Всего голосов:'
    ws['B7'] = total_votes
    ws['A7'].font = Font(bold=True)
    
    ws['A8'] = 'Дата создания:'
    ws['B8'] = poll_data['created_at'].split('T')[0] if poll_data.get('created_at') else 'N/A'
    ws['A8'].font = Font(bold=True)
    
    ws['A10'] = 'Вариант ответа'
    ws['B10'] = 'Голосов'
    ws['C10'] = 'Процент'
    
    for cell in ['A10', 'B10', 'C10']:
        ws[cell].fill = header_fill
        ws[cell].font = header_font
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
    
    row = 11
    for i, option in enumerate(poll_data['options']):
        count = stats[i]
        percentage = round((count / total_votes * 100) if total_votes > 0 else 0, 1)
        
        ws[f'A{row}'] = option
        ws[f'B{row}'] = count
        ws[f'C{row}'] = f'{percentage}%'
        
        ws[f'B{row}'].alignment = Alignment(horizontal='center')
        ws[f'C{row}'].alignment = Alignment(horizontal='center')
        
        row += 1
    
    chart = BarChart()
    chart.title = 'Распределение голосов'
    chart.x_axis.title = 'Варианты ответов'
    chart.y_axis.title = 'Количество голосов'
    
    data = Reference(ws, min_col=2, min_row=10, max_row=10+len(poll_data['options']))
    cats = Reference(ws, min_col=1, min_row=11, max_row=10+len(poll_data['options']))
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws.add_chart(chart, 'E10')
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()
    buffer.close()
    
    return excel_bytes

def handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    method: str = event.get('httpMethod', 'GET')
    
    if method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'GET, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type',
                'Access-Control-Max-Age': '86400'
            },
            'body': '',
            'isBase64Encoded': False
        }
    
    if method != 'GET':
        return {
            'statusCode': 405,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({'error': 'Method not allowed'}),
            'isBase64Encoded': False
        }
    
    params = event.get('queryStringParameters') or {}
    poll_id = params.get('poll_id')
    export_format = params.get('format', 'pdf').lower()
    
    if not poll_id:
        return {
            'statusCode': 400,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({'error': 'poll_id required'}),
            'isBase64Encoded': False
        }
    
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute('''
            SELECT id, target_audience, question, 
                   option1, option2, option3, option4, option5,
                   created_at
            FROM polls 
            WHERE id = %s AND is_active = true
        ''', (poll_id,))
        
        row = cur.fetchone()
        if not row:
            cur.close()
            conn.close()
            return {
                'statusCode': 404,
                'headers': {
                    'Content-Type': 'application/json',
                    'Access-Control-Allow-Origin': '*'
                },
                'body': json.dumps({'error': 'Poll not found'}),
                'isBase64Encoded': False
            }
        
        poll_data = {
            'id': row[0],
            'target_audience': row[1],
            'question': row[2],
            'options': [row[3], row[4], row[5], row[6], row[7]],
            'created_at': row[8].isoformat() if row[8] else None
        }
        
        stats = []
        for i in range(1, 6):
            cur.execute('''
                SELECT COUNT(*) 
                FROM poll_responses 
                WHERE poll_id = %s AND selected_option = %s
            ''', (poll_id, i))
            count = cur.fetchone()[0]
            stats.append(count)
        
        cur.close()
        conn.close()
        
        if export_format == 'pdf':
            file_bytes = generate_pdf(poll_data, stats)
            content_type = 'application/pdf'
            filename = f'poll_{poll_id}_results.pdf'
        elif export_format == 'excel':
            file_bytes = generate_excel(poll_data, stats)
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename = f'poll_{poll_id}_results.xlsx'
        else:
            return {
                'statusCode': 400,
                'headers': {
                    'Content-Type': 'application/json',
                    'Access-Control-Allow-Origin': '*'
                },
                'body': json.dumps({'error': 'Invalid format. Use pdf or excel'}),
                'isBase64Encoded': False
            }
        
        encoded_file = base64.b64encode(file_bytes).decode('utf-8')
        
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': content_type,
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Access-Control-Allow-Origin': '*'
            },
            'body': encoded_file,
            'isBase64Encoded': True
        }
        
    except (psycopg2.Error, ValueError, IOError) as e:
        return {
            'statusCode': 500,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({'error': str(e)}),
            'isBase64Encoded': False
        }